import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import messagebox
import openpyxl
import os

EMPLOYEE_FILE = "employees.xlsx"

def load_employees():
    if not os.path.exists(EMPLOYEE_FILE):
        return []
    wb = openpyxl.load_workbook(EMPLOYEE_FILE)
    ws = wb.active
    return [(str(row[0]), row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)]

def save_employees(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["เลขบัตรประชาชน", "ชื่อ-สกุล", "แผนก"])
    for row in data:
        ws.append([str(row[0]), row[1], row[2]])
    wb.save(EMPLOYEE_FILE)

def show_employee_list():
    window = tb.Toplevel() #สร้างหน้าต่างย่อย
    window.title("รายชื่อพนักงาน")
    window.geometry("700x450")

    employee_data = load_employees()

    def add_or_edit():
        cid = entry_id.get().strip()
        name = entry_name.get().strip()
        dept = entry_dept.get().strip()
        if not cid or not name or not dept:
            messagebox.showwarning("กรอกข้อมูลไม่ครบ", "กรุณากรอกให้ครบ")
            return
        if not cid.isdigit() or len(cid) != 13:
            messagebox.showerror("เลขบัตรไม่ถูกต้อง", "เลขบัตรประชาชนต้องเป็นตัวเลข 13 หลัก")
            return

        for row in employee_data:
            existing_cid, existing_name, _ = row
            if existing_cid != cid and existing_name == name:
                messagebox.showerror("ชื่อซ้ำ", f"ชื่อ '{name}' มีอยู่ในระบบแล้ว")
                return

        for i, row in enumerate(employee_data):
            if row[0] == cid:
                confirm = messagebox.askyesno(
                    "ยืนยันการแก้ไข",
                    f"มีเลขบัตรประชาชน '{cid}' อยู่แล้ว\nคุณต้องการแก้ไขข้อมูลนี้หรือไม่?"
                )
                if confirm:
                    employee_data[i] = (cid, name, dept)
                    update_table()
                    save_employees(employee_data)
                else:
                    clear_inputs()
                return

        employee_data.append((cid, name, dept))
        update_table()
        save_employees(employee_data)

    def delete_selected():
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("ไม่ได้เลือก", "กรุณาเลือกแถวก่อนลบ")
            return
        cid = str(tree.item(selected)["values"][0])
        nonlocal employee_data
        employee_data = [row for row in employee_data if str(row[0]) != cid]
        update_table()
        save_employees(employee_data)

    def update_table():
        tree.delete(*tree.get_children())
        for index, row in enumerate(employee_data):
            tag = "evenrow" if index % 2 == 0 else "oddrow"
            tree.insert("", "end", values=row, tags=(tag,))
        clear_inputs()

    def clear_inputs():
        entry_id.delete(0, END)
        entry_name.delete(0, END)
        entry_dept.delete(0, END)

    def on_row_select(event):
        selected = tree.selection()
        if selected:
            values = tree.item(selected)["values"]
            entry_id.delete(0, END)
            entry_id.insert(0, values[0])
            entry_name.delete(0, END)
            entry_name.insert(0, values[1])
            entry_dept.delete(0, END)
            entry_dept.insert(0, values[2])

    frame = tb.Frame(window, padding=10, borderwidth=1, relief="solid")
    frame.pack(fill=BOTH, expand=True)
    style = tb.Style()
    style.configure(
        "Custom.Treeview.Heading",
        font=("Segoe UI", 10, "bold")
    )
    style.configure(
        "Custom.Treeview",
        font=("Segoe UI", 10),
        rowheight=28,
        background="white",
        fieldbackground="white"
    )
    style.layout("Custom.Treeview", [("Treeview.treearea", {"sticky": "nswe"})])

    tree = tb.Treeview(
        frame,
        columns=("cid", "name", "dept"),
        show="headings",
        style="Custom.Treeview"
    )
    tree.heading("cid", text="เลขบัตรประชาชน")
    tree.heading("name", text="ชื่อ-สกุล")
    tree.heading("dept", text="แผนก")
    tree.column("cid", anchor="center", width=200, minwidth=150, stretch=False)
    tree.column("name", anchor="center", width=250, minwidth=200, stretch=False)
    tree.column("dept", anchor="center", width=100, stretch=False)

    tree.tag_configure("oddrow", background="#f8f9fa")
    tree.tag_configure("evenrow", background="#e9ecef")

    tree.grid(row=0, column=0, columnspan=5, sticky="nsew")
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    tree.bind("<<TreeviewSelect>>", on_row_select)

    scrollbar = tb.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    scrollbar.grid(row=0, column=5, sticky='ns')

    entry_id = tb.Entry(frame, width=20, bootstyle="info")
    entry_id.grid(row=1, column=0, padx=5, pady=10)

    entry_name = tb.Entry(frame, width=25, bootstyle="info")
    entry_name.grid(row=1, column=1, padx=5, pady=10)

    entry_dept = tb.Entry(frame, width=15, bootstyle="info")
    entry_dept.grid(row=1, column=2, padx=5, pady=10)

    btn_add = tb.Button(frame, text="Add/Edit", bootstyle="success", command=add_or_edit)
    btn_add.grid(row=1, column=3, padx=5, pady=10)

    btn_del = tb.Button(frame, text="Del", bootstyle="danger", command=delete_selected)
    btn_del.grid(row=1, column=4, padx=5, pady=10)

    update_table()
    window.mainloop()
