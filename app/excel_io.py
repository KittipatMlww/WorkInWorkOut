"""
excel_io.py  |  รวมทุกฟังก์ชัน read / write Excel (.xlsx) ในโฟลเดอร์ assets/
---------------------------------------------------------------------------
- load_employees()          → list[tuple]
- save_employees(data)      → เขียนไฟล์ employees.xlsx
- save_daily_log_row(row)   → เพิ่ม/อัปเดต 1 แถวใน daily_time_log.xlsx
- load_daily_logs()         → list[dict]  (อ่านทั้งไฟล์ log)
- logs_between(d1, d2)      → list[dict]  (กรองตามช่วงวันที่)
"""
from __future__ import annotations
import os
from datetime import datetime, date
from typing import List, Dict
import openpyxl

_ASSET_DIR = "assets"
EMPLOYEE_FILE   = os.path.join(_ASSET_DIR, "employees.xlsx")
DAILY_LOG_FILE  = os.path.join(_ASSET_DIR, "daily_time_log.xlsx")

# ---------- helper: make sure assets/ exists ----------
os.makedirs(_ASSET_DIR, exist_ok=True)

# ------------------------------------------------------------------------
# 1) พนักงาน
# ------------------------------------------------------------------------
def load_employees() -> List[tuple]:
    if not os.path.exists(EMPLOYEE_FILE):
        return []
    wb = openpyxl.load_workbook(EMPLOYEE_FILE, data_only=True)
    ws = wb.active
    data = [(str(c1), c2, c3) for c1, c2, c3, *_ in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return data

def save_employees(rows: List[tuple]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["เลขบัตรประชาชน", "ชื่อ-สกุล", "แผนก"])
    for pid, name, dept in rows:
        ws.append([str(pid), name, dept])
    wb.save(EMPLOYEE_FILE)

# ------------------------------------------------------------------------
# 2) Daily time log
# ------------------------------------------------------------------------
_LOG_HEADER = ["วันที่", "เวลาเข้า", "เวลาออก", "ชื่อสกุล",
               "เลขบัตรประชาชน", "แผนก"]

def _open_log_wb():
    """คืนค่า (wb, ws) หากไฟล์ยังไม่มีก็สร้างใหม่พร้อม header"""
    if os.path.exists(DAILY_LOG_FILE):
        wb = openpyxl.load_workbook(DAILY_LOG_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(_LOG_HEADER)
    return wb, ws

def save_daily_log_row(d: Dict[str, str]) -> None:
    """
    เพิ่ม/อัปเดต 1 แถว:
    - key ของแถว: (วันที่, เลขบัตรประชาชน)
    - ถ้าแถวเดิมยังไม่มีเวลาออก จะอัปเดต
    """
    wb, ws = _open_log_wb()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == d["วันที่"] and str(row[4]) == str(d["เลขบัตรประชาชน"]):
            # มีแถวนี้แล้ว
            if not row[2] and d["เวลาออก"]:
                ws.cell(row=idx, column=3, value=d["เวลาออก"])
            wb.save(DAILY_LOG_FILE)
            wb.close()
            return

    # ไม่พบแถวเดิม → เพิ่มใหม่
    ws.append([d[h] for h in _LOG_HEADER])
    wb.save(DAILY_LOG_FILE)
    wb.close()

# ------------------------------------------------------------------------
# 3) Utilities สำหรับรายงาน
# ------------------------------------------------------------------------
def load_daily_logs() -> List[Dict[str, str]]:
    """อ่านทุกแถวเป็น dict พร้อมคีย์ header"""
    logs = []
    if not os.path.exists(DAILY_LOG_FILE):
        return logs
    wb = openpyxl.load_workbook(DAILY_LOG_FILE, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        logs.append(dict(zip(_LOG_HEADER, row)))
    wb.close()
    return logs

def logs_between(d1: datetime, d2: datetime) -> List[Dict[str, str]]:
    """กรอง log ให้อยู่ในช่วง d1..d2 (inclusive)"""
    out = []
    for rec in load_daily_logs():
        try:
            rec_dt = datetime.strptime(rec["วันที่"], "%d/%m/%Y")
        except ValueError:
            # กรณีเซลล์เป็น datetime จริง
            if isinstance(rec["วันที่"], (datetime, date)):
                rec_dt = rec["วันที่"] if isinstance(rec["วันที่"], datetime) \
                         else datetime.combine(rec["วันที่"], datetime.min.time())
            else:
                continue
        if d1 <= rec_dt <= d2:
            out.append(rec)
    return out
